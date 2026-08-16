"""
DURATECH HTML 导出服务 - 接收勾选产品数据，生成格式化 Excel
由 html_builder.py 内嵌启动，html 页面通过 fetch POST 调用
"""
import json
import sys
import tempfile
from pathlib import Path
from datetime import datetime
from flask import Flask, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)  # 允许浏览器跨域访问（HTML 页面是 file:// 协议）

# 样式常量
FONT_NAME = "微软雅黑"
HEADER_FILL_COLOR = '2F5496'
LABEL_FILLS = {
    '潜力新品': 'E2EFDA',
    '标准爆品': 'FCE4D6',
    '头部爆品': 'F4B4B4',
}

# 类目名 → 友好 Sheet 名映射
CATEGORY_SHEET_NAMES = {
    "Arts, Crafts & Sewing": "ArtsCrafts",
    "Automotive": "Automotive",
    "Patio, Lawn & Garden": "PatioLawn",
    "Tools & Home Improvement": "ToolsHome",
    "ArtsCrafts": "ArtsCrafts",
    "PatioLawn": "PatioLawn",
    "ToolsHome": "ToolsHome",
}
EXPORT_FIELDS = [
    "产品类目", "ASIN", "品牌", "主图", "Item Name(标题)", "中文简介",
    "近30天销量(父体)", "月销售额", "零售价", "上架时间",
    "卖家信息", "评分", "评论数", "BSR排名", "变体数", "标签",
]


def build_excel(products, collect_type="new"):
    """用 openpyxl 生成格式化 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor=HEADER_FILL_COLOR)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name=FONT_NAME, size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    row_height = 80

    label_fills = {}
    for label_name, color in LABEL_FILLS.items():
        label_fills[label_name] = PatternFill('solid', fgColor=color)

    # 按类目分组
    category_groups = {}
    for p in products:
        cat = p.get('category', p.get('category_en', '其他'))
        if cat not in category_groups:
            category_groups[cat] = []
        category_groups[cat].append(p)

    # 每个类目一个 Sheet
    for cat_name, cat_products in category_groups.items():
        # 使用友好 Sheet 名，最多 31 字符
        sheet_name = CATEGORY_SHEET_NAMES.get(cat_name, cat_name)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 表头
        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        for row_idx, p in enumerate(cat_products, 2):
            label = p.get('__label__', p.get('label', ''))
            row_fill = label_fills.get(label)

            row_data = [
                cat_name,
                p.get('asin', ''),
                p.get('brand', ''),
                p.get('image', ''),
                p.get('title', ''),
                p.get('title_cn', p.get('_title_cn', '')),
                p.get('sales', ''),
                p.get('monthly_sales', ''),
                p.get('price', ''),
                p.get('available', ''),
                p.get('seller', ''),
                p.get('rating', ''),
                p.get('reviews', ''),
                p.get('bsr', ''),
                p.get('variants', ''),
                label,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

            ws.row_dimensions[row_idx].height = row_height

        # 列宽
        col_widths = [18, 14, 16, 35, 50, 50, 16, 16, 12, 14, 16, 8, 10, 12, 10, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 冻结首行 + 自动筛选
        ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(EXPORT_FIELDS))
        ws.auto_filter.ref = f"A1:{last_col}{len(cat_products) + 1}"

    # 如果没有产品，添加说明 Sheet
    if not wb.sheetnames:
        ws = wb.create_sheet("说明")
        ws['A1'] = "未选择任何产品"

    return wb


@app.route('/export', methods=['POST'])
def export():
    """接收产品 JSON，返回 Excel 文件"""
    try:
        data = request.get_json()
        products = data.get('products', [])
        collect_type = data.get('type', 'new')

        if not products:
            return {'error': '没有产品数据'}, 400

        wb = build_excel(products, collect_type)

        # 保存到临时文件
        date_str = datetime.now().strftime("%Y-%m-%d")
        prefix = "DuraTech_新品筛选" if collect_type == "new" else "DuraTech_爆品筛选"
        filename = f"{prefix}_{date_str}.xlsx"

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        return send_file(
            tmp.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/health', methods=['GET'])
def health():
    return {'status': 'ok'}


if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 58900
    print(f"[EXPORT] 导出服务启动: http://127.0.0.1:{port}")
    app.run(host='127.0.0.1', port=port, debug=False)
