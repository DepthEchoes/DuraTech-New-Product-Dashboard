"""
DURATECH JSON → Excel 转换工具
当 HTML 页面 fetch 失败时，浏览器会下载 .json 文件，用户运行此脚本即可生成格式化 Excel

用法: python json_to_excel.py <path-to-json>
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "微软雅黑"
EXPORT_FIELDS = [
    "产品类目", "ASIN", "品牌", "主图", "Item Name(标题)", "中文简介",
    "近30天销量(父体)", "月销售额", "零售价", "上架时间",
    "卖家信息", "评分", "评论数", "BSR排名", "变体数", "标签",
]

LABEL_FILLS = {
    '潜力新品': 'E2EFDA',
    '标准爆品': 'FCE4D6',
    '头部爆品': 'F4B4B4',
}

CATEGORY_SHEET_NAMES = {
    "Arts, Crafts & Sewing": "ArtsCrafts",
    "Automotive": "Automotive",
    "Patio, Lawn & Garden": "PatioLawn",
    "Tools & Home Improvement": "ToolsHome",
    "ArtsCrafts": "ArtsCrafts",
    "PatioLawn": "PatioLawn",
    "ToolsHome": "ToolsHome",
}


def build_excel(products, collect_type="new"):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name=FONT_NAME, size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    label_fills = {n: PatternFill('solid', fgColor=c) for n, c in LABEL_FILLS.items()}

    # 按类目分组
    category_groups = {}
    for p in products:
        cat = p.get('category', '其他')
        category_groups.setdefault(cat, []).append(p)

    for cat_name, cat_products in category_groups.items():
        ws = wb.create_sheet(title=CATEGORY_SHEET_NAMES.get(cat_name, cat_name)[:31])

        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, p in enumerate(cat_products, 2):
            label = p.get('__label__', p.get('label', ''))
            row_fill = label_fills.get(label)
            row_data = [
                cat_name,
                p.get('asin', ''),
                p.get('brand', ''),
                p.get('image', ''),
                p.get('title', ''),
                p.get('_title_cn', p.get('title_cn', '')),
                p.get('sales', ''),
                p.get('monthly_sales', ''),
                p.get('price', ''),
                p.get('available', ''),
                p.get('seller', ''),
                p.get('rating', ''),
                p.get('reviews', ''),
                p.get('bsr', ''),
                p.get('variants', ''),
                label,
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
            ws.row_dimensions[row_idx].height = 80

        col_widths = [18, 14, 16, 35, 50, 50, 16, 16, 12, 14, 16, 8, 10, 12, 10, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(EXPORT_FIELDS))
        ws.auto_filter.ref = f"A1:{last_col}{len(cat_products) + 1}"

    if not wb.sheetnames:
        ws = wb.create_sheet("说明")
        ws['A1'] = "未选择任何产品"

    return wb


def main():
    if len(sys.argv) < 2:
        print("用法: python json_to_excel.py <path-to-json>")
        print("示例: python json_to_excel.py DuraTech_新品筛选_2026-07-13.json")
        return 1

    json_path = Path(sys.argv[1])
    if not json_path.exists():
        print(f"错误: 找不到文件 {json_path}")
        return 1

    data = json.loads(json_path.read_text())
    products = data.get('products', [])
    collect_type = data.get('type', 'new')

    if not products:
        print("JSON 中没有产品数据")
        return 1

    print(f"读取到 {len(products)} 个产品")

    wb = build_excel(products, collect_type)

    # 输出到同目录
    out_path = json_path.with_suffix('.xlsx')
    wb.save(out_path)
    print(f"[OK] Excel 已生成: {out_path}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
