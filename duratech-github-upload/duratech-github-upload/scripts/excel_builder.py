"""
DURATECH Excel 输出生成器
新品+爆品 分开两个 Excel 文件
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from config import CATEGORIES, CATEGORY_SHORT_NAMES, EXPORT_FIELDS, WORKSPACE_OUTPUT, HOT_PRODUCT
from filter_rules import FilterEngine


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def translate_title(title):
    """用简单规则翻译标题（基于常见 Amazon 关键词映射）"""
    if not title:
        return ""
    # 常见产品词翻译映射
    terms = [
        ("Car Windshield Sunshade", "汽车挡风玻璃遮阳罩"),
        ("Windshield Sun Shade", "挡风玻璃遮阳板"),
        ("Windshield Sunshade", "挡风玻璃遮阳罩"),
        ("Tire Inflator", "轮胎充气器"),
        ("Portable Air Compressor", "便携式空气压缩机"),
        ("Air Compressor", "空气压缩机"),
        ("Car Mount", "车载支架"),
        ("Magsafe Car Mount", "磁吸车载支架"),
        ("Magnetic Phone Holder", "磁性手机支架"),
        ("Phone Holder", "手机支架"),
        ("Phone Holders", "手机支架"),
        ("Cabin Air Filter", "空调滤清器"),
        ("Air Filter", "空气过滤器"),
        ("Jump Starter", "应急启动电源"),
        ("Portable Lithium Jump Starter", "便携式锂电池应急启动电源"),
        ("Car Accessories", "汽车配件"),
        ("Automotive Sun Screen", "汽车遮阳帘"),
        ("Acrylic Paint", "丙烯颜料"),
        ("Acrylic Paint Markers", "丙烯颜料笔"),
        ("Paint Marker", "油漆笔"),
        ("Paint Pens", "油漆笔"),
        ("Flat Back Crystal Rhinestones", "平底水晶钻石"),
        ("Flat Back", "平底"),
        ("Rhinestones", "水钻"),
        ("Crystal Rhinestones", "水晶钻石"),
        ("Essential Oils", "精油"),
        ("Tea Tree Oil", "茶树精油"),
        ("Flying Insect Trap", "飞虫捕捉器"),
        ("Insect Trap", "昆虫捕捉器"),
        ("Ant Killer", "蚂蚁药"),
        ("Ant Killer Bait Stations", "蚂蚁诱饵站"),
        ("Water Filter", "滤水器"),
        ("Standard Water Filter", "标准滤水器"),
        ("Replacement Water Filter", "替换滤水器"),
        ("Painter's Tape", "美纹纸胶带"),
        ("Spray Paint", "喷漆"),
        ("Cordless", "无绳"),
        ("Rechargeable", "可充电"),
        ("Waterproof", "防水"),
        ("Stainless Steel", "不锈钢"),
        ("Aluminum", "铝"),
        ("Copper", "铜"),
        ("Professional", "专业级"),
        ("Heavy Duty", "重型"),
        ("Premium", "优质"),
        ("Ultra", "超"),
        ("Blocks UV Rays", "阻挡紫外线"),
        ("Keeps Interior Cool", "保持车内凉爽"),
        ("Keeps Car Interior Cool", "保持车内凉爽"),
        ("Blocks 99% Heat", "阻挡99%热量"),
        ("Universal Fit", "通用适配"),
        ("for Car", "汽车用"),
        ("for Cars", "汽车用"),
        ("for Trucks", "卡车用"),
        ("for SUVs", "SUV用"),
        ("for Toyota", "丰田"),
        ("for Honda", "本田"),
        ("for iPhone", "苹果手机"),
        ("for Men", "男士"),
        ("for Women", "女士"),
        ("Set", "套装"),
        ("Kit", "套件"),
        ("with", "带"),
        ("and", "和"),
    ]

    result = title
    for en, cn in terms:
        result = result.replace(en, cn)

    # 如果翻译后和原文一样（没有匹配），返回通用描述
    if result == title:
        # 提取前 60 个字符做基础描述
        result = f"[{title[:50]}]"

    return result


class ExcelBuilder:
    FONT_NAME = "微软雅黑"
    HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='2F5496')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_FONT = Font(name=FONT_NAME, size=10)
    DATA_ALIGN = Alignment(vertical='center', wrap_text=True)
    LABEL_FILLS = {
        '潜力新品': PatternFill('solid', fgColor='E2EFDA'),
        '标准爆品': PatternFill('solid', fgColor='FCE4D6'),
        '头部爆品': PatternFill('solid', fgColor='F4B4B4'),
    }
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ROW_HEIGHT = 80

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def _map_fields(self, p):
        title = p.get("title", "")
        return {
            "产品类目": p.get("category", ""),
            "ASIN": p.get("asin", ""),
            "品牌": p.get("brand", ""),
            "主图": p.get("image", ""),
            "Item Name(标题)": title,
            "中文简介": translate_title(title),
            "近30天销量(父体)": p.get("sales", ""),
            "月销售额": p.get("monthly_sales", ""),
            "零售价": p.get("price", ""),
            "上架时间": p.get("available", ""),
            "卖家信息": p.get("seller", ""),
            "评分": p.get("rating", ""),
            "评论数": p.get("reviews", ""),
            "BSR排名": p.get("bsr", ""),
            "变体数": p.get("variants", ""),
            "标签": p.get("__label__", ""),
        }

    def create_sheet(self, sheet_name, products):
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = self.wb.create_sheet(title=sheet_name)

        # 表头
        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.BORDER

        # 数据
        for row_idx, p in enumerate(products, 2):
            mapped = self._map_fields(p)
            label = mapped.get("标签", "")
            row_fill = self.LABEL_FILLS.get(label, None)

            for col_idx, field in enumerate(EXPORT_FIELDS, 1):
                value = mapped.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGN
                cell.border = self.BORDER
                if row_fill:
                    cell.fill = row_fill

            ws.row_dimensions[row_idx].height = self.ROW_HEIGHT

        # 列宽 (16列)
        col_widths = {
            'A': 18, 'B': 14, 'C': 16, 'D': 35, 'E': 50,
            'F': 50, 'G': 16, 'H': 16, 'I': 12, 'J': 14,
            'K': 16, 'L': 8, 'M': 10, 'N': 12, 'O': 10, 'P': 16,
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(EXPORT_FIELDS))
        ws.auto_filter.ref = f"A1:{last_col}{len(products)+1}"

        return ws

    def build(self, category_data, output_path):
        for cat_en in CATEGORIES:
            short = CATEGORY_SHORT_NAMES.get(cat_en, cat_en)[:10]
            products = category_data.get(cat_en, [])
            if products:
                self.create_sheet(short, products)
                log(f"  [{short}] {len(products)} 条")

        if not self.wb.sheetnames:
            ws = self.wb.create_sheet("说明")
            ws['A1'] = "本次未采集到符合条件的产品数据"

        self.wb.save(output_path)
        log(f"[OK] Excel 已生成: {output_path}")
        return output_path


def assign_labels(products, collect_type):
    if collect_type == "new":
        for p in products:
            p['__label__'] = '潜力新品'
    elif collect_type == "hot":
        for p in products:
            sales = int(p.get('sales', 0)) if p.get('sales', '').isdigit() else 0
            if sales >= HOT_PRODUCT['head_min']:
                p['__label__'] = '头部爆品'
            elif sales >= HOT_PRODUCT['standard_min']:
                p['__label__'] = '标准爆品'
            else:
                p['__label__'] = '爆品'
    return products


def build_from_json(json_path, collect_type, output_path=None):
    if output_path is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        from config import COLLECT_MODES
        prefix = COLLECT_MODES.get(collect_type, {}).get('output_prefix', 'DuraTech')
        output_path = Path(WORKSPACE_OUTPUT) / f"{prefix}_{date_str}.xlsx"

    engine = FilterEngine()
    raw = json.loads(Path(json_path).read_text())
    category_data = {}

    for cat, products in raw.items():
        filtered = engine.deduplicate(products, key='asin')
        filtered = assign_labels(filtered, collect_type)
        category_data[cat] = filtered
        log(f"  [{collect_type}] {cat}: {len(filtered)} 条")

    builder = ExcelBuilder()
    return builder.build(category_data, output_path)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--json', '-j', required=True, help='原始 JSON 数据文件')
    parser.add_argument('--type', '-t', choices=['new', 'hot'], required=True, help='new/hot')
    parser.add_argument('--output', '-o', help='输出 Excel 路径')
    args = parser.parse_args()

    build_from_json(
        json_path=args.json,
        collect_type=args.type,
        output_path=args.output,
    )
